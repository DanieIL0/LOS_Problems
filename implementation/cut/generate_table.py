from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import os
import pytz
import re

def collect_segment_info(
    segment_info_list,
    segment_info,
    los_issue_duration,
    segment_index,
    log_steps,
    log_step_description,
    los_issue_start_time,
    trial_number,
    pretrial,
    trial_type
):
    """
    Collects information about a video segment and appends it to the segment_info_list if:
    - The segment's timestamp is in ascending order (with a 0.5-second buffer).
    - The combination of 'Original Videos' and 'Segment' number is unique.

    Parameters:
        segment_info_list (list): List to store segment information dictionaries.
        segment_info (dict): Information about the current segment.
        los_issue_duration (float): Duration of the LOS issue.
        segment_index (int): Index of the segment.
        log_steps (list): List of parsed log steps.
        log_step_description (str): Description of the log step.
        los_issue_start_time (float): Start time of the LOS issue.
        trial_number (str): The trial number extracted from the directory name.
        pretrial (bool): Indicates if it's a pretrial.
        trial_type (str): The trial type extracted from the directory name.
    """
    trial_number = str(int(trial_number)) if trial_number.isdigit() else trial_number
    origin_videos_info = []
    for vid_file, vid_start, vid_end in segment_info['video_inputs']:
        origin_videos_info.append({
            'vid_file': vid_file,
            'vid_start': vid_start,
            'vid_end': vid_end
        })

    patterns_to_replace = ['1-*', '3-*', '4-*']
    pattern_str = '_(?:' + '|'.join(map(re.escape, patterns_to_replace)) + ')'
    origin_videos_set = set()
    for info in origin_videos_info:
        vid_file = info['vid_file']
        vid_file_cleaned = re.sub(pattern_str, '*', vid_file)
        origin_videos_set.add(vid_file_cleaned)
    origin_videos_str = '+'.join(sorted(origin_videos_set))

    segment_number = segment_index + 1
    local_tz = pytz.timezone('Europe/Berlin')
    segment_start_datetime = datetime.fromtimestamp(segment_info['segment_start_time'], tz=pytz.utc)
    segment_start_datetime = segment_start_datetime.astimezone(local_tz)
    day = segment_start_datetime.strftime('%d/%m/%Y')

    length_secs = los_issue_duration

    if pretrial or log_steps is None or log_step_description is None:
        performed_step = ''
        step_length_mmss = ''
    else:
        performed_step = log_step_description if log_step_description else 'NaN'

        if log_steps and log_step_description:
            for step in log_steps:
                if step['description'] == log_step_description:
                    step_length_secs = step['end_time'] - step['start_time']
                    minutes = int(step_length_secs // 60)
                    seconds = int(step_length_secs % 60)
                    step_length_mmss = f"{minutes}:{seconds:02d}"
                    break
            else:
                step_length_mmss = 'NaN'
        else:
            step_length_mmss = 'NaN'

    los_issue_start_datetime = datetime.fromtimestamp(los_issue_start_time, tz=pytz.utc)
    los_issue_start_datetime = los_issue_start_datetime.astimezone(local_tz)
    los_issue_start_time_str = los_issue_start_datetime.strftime('%H:%M:%S')

    segment_data = {
        'Pretrial': pretrial,
        'Trial': trial_type,
        'Trial Number': trial_number,
        'Original Videos': origin_videos_str,
        'Segment': segment_number,
        'Day': day,
        'LOS Issue Start Time': los_issue_start_time_str,
        'Length (secs)': f"{length_secs:.2f}",
        'Performed Step': performed_step,
        'Length of step (mm:ss)': step_length_mmss,
        'Reason': ''
    }
    date_time_str = f"{segment_data['Day']}_{segment_data['LOS Issue Start Time']}"

    duplicate_exists = any(
        (seg['Segment'] == segment_data['Segment']) and
        (seg['Trial Number'] == segment_data['Trial Number']) and
        (seg['Length (secs)'] == segment_data['Length (secs)']) and
        (f"{seg['Day']}_{seg['LOS Issue Start Time']}" == date_time_str)
        for seg in segment_info_list
    )

    if duplicate_exists:
        return
    segment_info_list.append(segment_data)

def generate_excel_table(segment_info_list, excel_output_path):
    """
    Generates an Excel table from the segment information list.
    If the Excel file already exists, it appends the new data to it.

    Parameters:
        segment_info_list (list): List of dictionaries containing segment information.
        excel_output_path (str): Path where the Excel file will be saved.
    """
    new_df = pd.DataFrame(segment_info_list)

    if os.path.exists(excel_output_path):
        existing_df = pd.read_excel(excel_output_path, engine='openpyxl')
        df = pd.concat([existing_df, new_df], ignore_index=True)
    else:
        df = new_df
    df.to_excel(excel_output_path, index=False, engine='openpyxl')

    wb = load_workbook(excel_output_path)
    ws = wb.active

    column_letters = {cell.value: cell.column_letter for cell in ws[1]}

    if 'Original Videos' in column_letters:
        origin_video_column = column_letters['Original Videos']
        ws.column_dimensions[origin_video_column].width = 45

    if 'Reason' in column_letters:
        reason_column = column_letters['Reason']
        ws.column_dimensions[reason_column].width = 60

    if 'Performed Step' in column_letters:
        performed_step_column = column_letters['Performed Step']
        ws.column_dimensions[performed_step_column].width = 45

    wb.save(excel_output_path)